#!/usr/bin/env python3
"""
create-pack-xlsx.py - Extract investor pack sheets from a full XLSX workbook.

Creates a curated "pack" workbook containing only investor-facing sheets
with proper print areas and page setup for PDF export.

Usage:
    python3 scripts/create-pack-xlsx.py <source.xlsx> <output_pack.xlsx> [--config path/to/config.json]

Example:
    python3 scripts/create-pack-xlsx.py artifacts/golden.xlsx artifacts/golden_pack.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    from copy import copy
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# Default print boundaries when source sheet lacks print area
# These are conservative estimates based on typical IND_ACQ content
DEFAULT_PRINT_COLUMNS = {
    "Investment Summary": "M",      # ~13 columns
    "Returns Summary": "G",         # ~7 columns
    "Error Check": "F",             # ~6 columns
    "Model Outputs": "H",           # ~8 columns
    "Annual CF": "N",               # ~14 columns (years 0-10 + labels)
    "Assumptions": "F",             # ~6 columns
    "Rent Roll": "L",               # ~12 columns
    "Renovation Budget": "E",       # ~5 columns
    "Monthly CF": "BM",             # ~65 columns (60 months + headers)
}

DEFAULT_PRINT_ROWS = {
    "Investment Summary": 50,
    "Returns Summary": 40,
    "Error Check": 30,
    "Model Outputs": 60,
    "Annual CF": 45,
    "Assumptions": 80,
    "Rent Roll": 50,
    "Renovation Budget": 40,
    "Monthly CF": 60,
}


def load_pack_config(config_path: Path) -> dict:
    """Load the pack export configuration JSON."""
    with open(config_path, "r") as f:
        return json.load(f)


def get_print_area(sheet, sheet_name: str) -> str:
    """
    Get print area for a sheet.
    Uses existing print area if defined, otherwise applies defaults.
    """
    if sheet.print_area:
        # Sheet has defined print area - use it
        return sheet.print_area

    # Apply default boundaries
    max_col = DEFAULT_PRINT_COLUMNS.get(sheet_name, "Z")
    max_row = DEFAULT_PRINT_ROWS.get(sheet_name, 100)
    return f"A1:{max_col}{max_row}"


def apply_page_setup(dest_sheet, config_sheet: dict, source_sheet):
    """Apply page setup from config and source sheet."""
    orientation = config_sheet.get("orientation", "portrait")

    # Copy page setup from source if available
    if source_sheet.page_setup:
        dest_sheet.page_setup.paperSize = source_sheet.page_setup.paperSize
        dest_sheet.page_setup.scale = source_sheet.page_setup.scale
        dest_sheet.page_setup.fitToPage = source_sheet.page_setup.fitToPage
        dest_sheet.page_setup.fitToWidth = source_sheet.page_setup.fitToWidth
        dest_sheet.page_setup.fitToHeight = source_sheet.page_setup.fitToHeight
    else:
        # Apply defaults for fit-to-page
        dest_sheet.page_setup.fitToPage = True
        dest_sheet.page_setup.fitToWidth = 1
        dest_sheet.page_setup.fitToHeight = 0  # 0 = auto height

    # Override orientation from config
    dest_sheet.page_setup.orientation = orientation

    # Apply margins from config print_settings
    dest_sheet.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.5,
        bottom=0.5,
        header=0.3,
        footer=0.3
    )


def copy_sheet_content(source_sheet, dest_sheet):
    """Copy cell values, styles, and merged cells from source to dest."""
    # Copy merged cells first
    for merged_range in source_sheet.merged_cells.ranges:
        dest_sheet.merge_cells(str(merged_range))

    # Copy cell data and styles
    for row in source_sheet.iter_rows():
        for cell in row:
            # Skip MergedCell objects (only top-left cell of merge has data)
            if isinstance(cell, MergedCell):
                continue

            dest_cell = dest_sheet.cell(row=cell.row, column=cell.column)
            dest_cell.value = cell.value

            if cell.has_style:
                dest_cell.font = copy(cell.font)
                dest_cell.border = copy(cell.border)
                dest_cell.fill = copy(cell.fill)
                dest_cell.number_format = cell.number_format
                dest_cell.alignment = copy(cell.alignment)
                dest_cell.protection = copy(cell.protection)

    # Copy column widths
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        dest_sheet.column_dimensions[col_letter].width = col_dim.width
        dest_sheet.column_dimensions[col_letter].hidden = col_dim.hidden

    # Copy row heights
    for row_num, row_dim in source_sheet.row_dimensions.items():
        dest_sheet.row_dimensions[row_num].height = row_dim.height
        dest_sheet.row_dimensions[row_num].hidden = row_dim.hidden


def create_pack_workbook(source_path: Path, output_path: Path, config: dict) -> dict:
    """
    Create a pack workbook with only investor-facing sheets.

    Returns dict with status info.
    """
    # Load source workbook
    print(f"Loading source workbook: {source_path}")
    source_wb = load_workbook(source_path, data_only=False)
    source_sheets = source_wb.sheetnames
    print(f"  Source sheets: {source_sheets}")

    # Get pack sheets from config (sorted by order)
    pack_sheets = sorted(config.get("pack_sheets", []), key=lambda x: x.get("order", 999))

    # Create new workbook
    dest_wb = Workbook()
    # Remove default sheet
    default_sheet = dest_wb.active

    sheets_copied = []
    sheets_missing = []
    total_expected_pages = 0

    for i, sheet_config in enumerate(pack_sheets):
        sheet_name = sheet_config["name"]
        expected_pages = sheet_config.get("expected_pages", 1)

        if sheet_name not in source_sheets:
            print(f"  WARNING: Sheet '{sheet_name}' not found in source")
            sheets_missing.append(sheet_name)
            continue

        source_sheet = source_wb[sheet_name]

        # Create destination sheet
        if i == 0:
            # Rename the default sheet for first one
            default_sheet.title = sheet_name
            dest_sheet = default_sheet
        else:
            dest_sheet = dest_wb.create_sheet(title=sheet_name)

        print(f"  Copying sheet: {sheet_name} (expected {expected_pages} pages)")

        # Copy content
        copy_sheet_content(source_sheet, dest_sheet)

        # Set print area
        print_area = get_print_area(source_sheet, sheet_name)
        dest_sheet.print_area = print_area
        print(f"    Print area: {print_area}")

        # Apply page setup
        apply_page_setup(dest_sheet, sheet_config, source_sheet)

        sheets_copied.append(sheet_name)
        total_expected_pages += expected_pages

    # Save output workbook
    print(f"Saving pack workbook: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dest_wb.save(output_path)

    result = {
        "success": True,
        "source": str(source_path),
        "output": str(output_path),
        "sheets_copied": sheets_copied,
        "sheets_missing": sheets_missing,
        "expected_pages": total_expected_pages,
        "page_tolerance": config.get("page_tolerance", 2),
    }

    print(f"\nPack created successfully:")
    print(f"  Sheets copied: {len(sheets_copied)}")
    print(f"  Expected pages: {total_expected_pages} (±{result['page_tolerance']})")

    if sheets_missing:
        print(f"  WARNING: Missing sheets: {sheets_missing}")

    return result


def main():
    parser = argparse.ArgumentParser(
        description="Create investor pack XLSX from full workbook"
    )
    parser.add_argument("source", help="Source XLSX file path")
    parser.add_argument("output", help="Output pack XLSX file path")
    parser.add_argument(
        "--config",
        default="docs/IND_ACQ_PACK_EXPORT.json",
        help="Pack export config JSON (default: docs/IND_ACQ_PACK_EXPORT.json)"
    )
    parser.add_argument(
        "--json",
        action="store_true",
        help="Output result as JSON"
    )

    args = parser.parse_args()

    source_path = Path(args.source)
    output_path = Path(args.output)
    config_path = Path(args.config)

    # Validate inputs
    if not source_path.exists():
        print(f"ERROR: Source file not found: {source_path}", file=sys.stderr)
        sys.exit(1)

    if not config_path.exists():
        print(f"ERROR: Config file not found: {config_path}", file=sys.stderr)
        sys.exit(1)

    # Load config
    config = load_pack_config(config_path)

    # Create pack workbook
    try:
        result = create_pack_workbook(source_path, output_path, config)

        if args.json:
            print(json.dumps(result, indent=2))

        sys.exit(0 if result["success"] else 1)

    except Exception as e:
        print(f"ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
